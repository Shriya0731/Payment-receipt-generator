import openpyxl as op
from datetime import datetime, date
from pyinvoice.models import InvoiceInfo, ServiceProviderInfo, ClientInfo, Item, Transaction
from pyinvoice.templates import SimpleInvoice
import pip

doc = SimpleInvoice('invoicenew.pdf')
# data which we are going to display as tables

workbook = op.load_workbook("productlist.xlsx")
sheet = workbook.active
wb2 = op.load_workbook("order.xlsx")
sheet2 = wb2.active

index=1
print(sheet2.max_row)
for j in range(2, sheet2.max_row+1):
    order_cell = sheet2.cell(row=j, column =1)
    for i in range(2,sheet.max_row + 1):
        my_cell = sheet.cell(row=i, column=1)

        if(my_cell.value==order_cell.value):
            #item name
            price= sheet.cell(row=i,column=2).value
            qty =  sheet2.cell(row=j, column=2).value
            doc.add_item(Item(order_cell.value, "Home Automation", qty, price))

doc.invoice_info = InvoiceInfo(1023, datetime.now(), datetime.now())  # Invoice info, optional
# Service Provider Info, optional
doc.service_provider_info = ServiceProviderInfo(name= "Kashtronica",
    street='Sinhagad Road',
    city='Pune',
    state='Maharashtra',
    country='India',
    post_code='411051',
    vat_tax_number='Vat/Tax number'
)

# Client info
doc.client_info = ClientInfo(email='client@example.com')

# Tax rate
doc.set_item_tax_rate(20)  # 20%
# Optional
doc.set_bottom_tip("Visit us at \n www.kashtronica.in<br />Don't hesitate to contact us for any questions.")

doc.finish()
